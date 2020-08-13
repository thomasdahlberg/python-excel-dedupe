
import openpyxl
import os
import sys

print('Opening workbook')

file_arg = sys.argv[1]
dedupe_arg = sys.argv[2]
sort_arg = sys.argv[3]

dedupe_criteria = dedupe_arg.split(',')


os.chdir('./workdocs')
wb = openpyxl.load_workbook(f"{file_arg}.xlsx")
sheet = wb['sheet1']
header_values = []
row_keys = range(2,sheet.max_row)
document = []
deduped_document = []
unique_values = []

for i in range(1, sheet.max_column + 1):
    header_values.append(sheet.cell(row=1, column=i).value)

print('Headers: ', header_values)
print('Row Keys: ', row_keys)

filter_cols = []
for i in range(0, len(dedupe_criteria)):
    filter_cols.append(header_values[int(dedupe_criteria[i])])
print('Dedupe Criteria: ', filter_cols)

sort_col = header_values[int(sort_arg)]
print('Sorting Criteria: ', sort_col)

# Deduplicating
for key in row_keys:
    row_dict = {}
    for i in range(1, sheet.max_column +1):
        row_dict[header_values[i-1]] = sheet.cell(row=key, column=i).value
    document.append(row_dict)

sorted_doc = sorted(document, key=lambda x: str(x[sort_col]))

for d in sorted_doc:
    criteria = []
    for i in range(0,len(filter_cols)):
        criteria.append(d[filter_cols[i]])
    if criteria not in unique_values:
        unique_values.append(criteria)
        deduped_document.append(d)

print('Original length: ', len(sorted_doc))
print('Deduped length: ', len(deduped_document))


# Writing to new tab
if 'Deduped' in wb.sheetnames:
    wb.remove(wb['Deduped'])

wb.create_sheet(index=1, title='Deduped')
deduped_sheet = wb['Deduped']
deduped_sheet.append(header_values)

for i in range(1, len(deduped_document)):
    deduped_sheet.append(list(deduped_document[i].values()))

wb.save(filename=f"{file_arg}.xlsx")

print("Process Complete!")